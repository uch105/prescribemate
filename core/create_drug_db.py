import os
import sys
import django
from decouple import config

project_root = config("PROJECT_ROOT")
sys.path.append(project_root)

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'prescribemate.settings')
django.setup()

from core.models import Drug
from openpyxl import load_workbook

def retype(s):
    if 'Tablet' in s:
        re_type = 'Tab.'
    elif 'Capsule' in s:
        re_type = 'Cap.'
    elif 'Injection' in s:
        re_type = 'Inj.'
    elif 'Infusion' in s:
        re_type = 'Inf.'
    elif 'Suppository' in s:
        re_type = 'Supp.'
    elif 'Syrup' in s:
        re_type = 'Syp.'
    else:
        re_type = s

    return re_type 

def rebrand(type,brand,strength):
    re_type = retype(type)
    if re_type==type:
        rebrand_tuple = (brand,re_type,'(',strength,')')
    else:
        rebrand_tuple = (re_type,brand,'(',strength,')')
    return ' '.join(rebrand_tuple)

path = '/home/uch/Downloads/prescribematedev/devfiles/drugs.xlsx'

workbook = load_workbook(path)
sheet = workbook.active

counter = 0

for row in range(1,sheet.max_row+1):
    manufacturer = sheet[f'A{row}'].value
    brand_name = sheet[f'B{row}'].value
    generic = sheet[f'C{row}'].value
    drugs_dose = sheet[f'D{row}'].value
    drugs_type = sheet[f'E{row}'].value
    price = sheet[f'F{row}'].value
    drugs_for = sheet[f'G{row}'].value
    drug_id = sheet[f'H{row}'].value

    try:
        Drug.objects.update_or_create(
            drugs_id = drug_id,
            brand = rebrand(type=drugs_type,brand=brand_name,strength=drugs_dose),
            generic = generic,
            manufacturer = manufacturer,
            strength = drugs_dose,
            applicable_for = drugs_for,
            indication = '',
            contraindication = '',
            side_effect = '',
        )
        counter+=1
    except:
        print(drug_id+' is a duplicate')

    if counter%500==0:
        print(f'{counter} drugs added!')

print(str(counter) + "Drugs imported successfully!")